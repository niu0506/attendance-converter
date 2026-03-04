import sys
import re
import warnings
import traceback
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：缺少依赖库 openpyxl，请执行：pip install openpyxl")
    input("\n按回车键退出...")
    sys.exit(1)

HEADER_FILL = PatternFill("solid", fgColor="646A73")
HEADER_FONT = Font(name="宋体", bold=True, color="FFFFFF", size=14)
SUB_FONT    = Font(name="宋体", bold=False, color="FFFFFF", size=14)
DATA_FONT   = Font(name="宋体", size=14)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_HEIGHT  = 20.0
COL_WIDTH   = 14.0

THIN   = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

KEEP_DEPTS = {"办公室", "多元化", "铁护办", "网格化", "信息化", "综治中心主任"}

TARGET_COLS = ["姓名", "部门", "日期",
               "上班1打卡时间", "下班1打卡时间",
               "上班2打卡时间", "下班2打卡时间"]

TIME_COLS = {"上班1打卡时间", "下班1打卡时间", "上班2打卡时间", "下班2打卡时间"}

GROUP_MAP = {
    "姓名": "基本信息", "部门": "基本信息", "日期": "基本信息",
    "上班1打卡时间": "打卡信息", "下班1打卡时间": "打卡信息",
    "上班2打卡时间": "打卡信息", "下班2打卡时间": "打卡信息",
}


def normalize_time(val):
    if val is None or str(val).strip() in ("", "-", " "):
        return "-"
    s = str(val).strip()
    if re.match(r"^\d{1,2}:\d{2}$", s):
        h, m = s.split(":")
        return f"{int(h):02d}:{m}"
    try:
        f = float(s)
        if 0 <= f < 1:
            total_minutes = round(f * 24 * 60)
            h, m = divmod(total_minutes, 60)
            return f"{h:02d}:{m:02d}"
    except ValueError:
        pass
    return s


def apply_style(cell, font, fill=None, alignment=CENTER, border=BORDER):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def convert(src_path: str, dst_path: str):
    import os

    if not os.path.exists(src_path):
        raise FileNotFoundError(f"源文件不存在：{src_path}")

    try:
        with open(src_path, "rb"):
            pass
    except PermissionError:
        raise PermissionError(f"文件被占用，请关闭 Excel 后重试：{src_path}")

    try:
        wb_src = load_workbook(src_path)
    except Exception as e:
        raise RuntimeError(f"无法读取文件（可能已损坏或格式不支持）：{e}")

    ws_src = wb_src.active

    if ws_src.max_row < 2:
        raise ValueError("文件内容不足：至少需要2行（第2行为表头）")

    header_row = [cell.value for cell in ws_src[2]]

    col_index = {}
    missing_cols = []
    for col_name in TARGET_COLS:
        try:
            col_index[col_name] = header_row.index(col_name)
        except ValueError:
            print(f"  警告：找不到列 '{col_name}'，将填充空值")
            col_index[col_name] = None
            missing_cols.append(col_name)

    if "部门" in missing_cols:
        raise ValueError("缺少关键列'部门'，无法过滤数据，请检查源文件格式")

    dept_col_idx = col_index.get("部门")

    data_rows = []
    for row_num, row in enumerate(ws_src.iter_rows(min_row=3, values_only=True), start=3):
        try:
            dept = row[dept_col_idx] if dept_col_idx is not None else None
            if dept not in KEEP_DEPTS:
                continue
            extracted = []
            for col_name in TARGET_COLS:
                idx = col_index[col_name]
                val = row[idx] if idx is not None else None
                if col_name in TIME_COLS:
                    val = normalize_time(val)
                extracted.append(val)
            data_rows.append(extracted)
        except Exception as e:
            print(f"  警告：第 {row_num} 行数据异常，已跳过（{e}）")
            continue

    if not data_rows:
        print(f"  警告：未找到匹配部门的数据，输出文件将为空表")

    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = ws_src.title

    groups = {}
    for i, col_name in enumerate(TARGET_COLS):
        grp = GROUP_MAP[col_name]
        if grp not in groups:
            groups[grp] = [i + 1, i + 1]
        else:
            groups[grp][1] = i + 1

    for grp, (start_col, end_col) in groups.items():
        cell = ws_dst.cell(row=1, column=start_col, value=grp)
        apply_style(cell, HEADER_FONT, HEADER_FILL)
        if start_col != end_col:
            ws_dst.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1,   end_column=end_col
            )
        for c in range(start_col + 1, end_col + 1):
            apply_style(ws_dst.cell(row=1, column=c), HEADER_FONT, HEADER_FILL)

    for i, col_name in enumerate(TARGET_COLS, start=1):
        cell = ws_dst.cell(row=2, column=i, value=col_name)
        apply_style(cell, SUB_FONT, HEADER_FILL)

    for r_idx, row_data in enumerate(data_rows, start=3):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws_dst.cell(row=r_idx, column=c_idx, value=value)
            apply_style(cell, DATA_FONT, alignment=CENTER)

    for r in range(1, 3 + len(data_rows)):
        ws_dst.row_dimensions[r].height = ROW_HEIGHT

    for i in range(1, len(TARGET_COLS) + 1):
        ws_dst.column_dimensions[get_column_letter(i)].width = COL_WIDTH

    ws_dst.page_setup.orientation = "landscape"

    out_dir = os.path.dirname(dst_path)
    if out_dir and not os.access(out_dir, os.W_OK):
        raise PermissionError(f"无权限写入目录：{out_dir}")

    try:
        wb_dst.save(dst_path)
    except PermissionError:
        raise PermissionError(f"输出文件被占用，请关闭后重试：{dst_path}")
    except Exception as e:
        raise RuntimeError(f"保存文件失败：{e}")


if __name__ == "__main__":
    import glob, os

    try:
        if len(sys.argv) == 3:
            src, dst = sys.argv[1], sys.argv[2]
            out_dir = os.path.dirname(dst)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            convert(src, dst)
            print(f"完成：{src} -> {dst}")
        else:
            matches = glob.glob("每日统计表*.xlsx")
            if not matches:
                print("错误：当前目录下未找到以'每日统计表'开头的xlsx文件")
                print(f"当前目录：{os.path.abspath('.')}")
                print("请将程序与源文件放在同一目录下运行")
            else:
                out_dir = "output"
                os.makedirs(out_dir, exist_ok=True)

                print(f"共找到 {len(matches)} 个文件，开始批量处理...\n")
                ok, fail = 0, 0
                for src in sorted(matches):
                    dst = os.path.join(out_dir, os.path.basename(src))
                    try:
                        convert(src, dst)
                        print(f"  ✓ {src} -> {dst}")
                        ok += 1
                    except FileNotFoundError as e:
                        print(f"  ✗ 文件不存在：{e}")
                        fail += 1
                    except PermissionError as e:
                        print(f"  ✗ 权限错误：{e}")
                        fail += 1
                    except ValueError as e:
                        print(f"  ✗ 数据格式错误：{e}")
                        fail += 1
                    except RuntimeError as e:
                        print(f"  ✗ 处理失败：{e}")
                        fail += 1
                    except Exception as e:
                        print(f"  ✗ 未知错误：{e}")
                        fail += 1

                print(f"\n完成：成功 {ok} 个，失败 {fail} 个")

    except KeyboardInterrupt:
        print("\n已取消")
    except Exception as e:
        print(f"\n程序异常：{e}")
        traceback.print_exc()

    if len(sys.argv) == 1:
        input("\n按回车键退出...")